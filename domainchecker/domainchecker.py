import requests
import json
import openpyxl

# Enter your VirusTotal API key here
# api_key = "6d20f0d4c9d8b0859dfea71e5937c6372e72813a617e10c582ad78215a74a017"

apifile = open('api.txt', 'r')
api_key = apifile.readline()

# Enter the path to the file containing the list of domains
input_file = "input.txt"
output_file = "output.xlsx"


# Define the VirusTotal API endpoint
url_domain = "https://www.virustotal.com/api/v3/domains/{}"

# Read the list of domains from the file
with open(input_file, "r") as f:
    domains = f.read().splitlines()

# Create a new Excel workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active

# Loop through each domain and check its reputation
for i, domain in enumerate(domains, start=1):
    headers = {
        "x-apikey": api_key,
        "Accept": "application/json"
    }
    r = requests.get(url_domain.format(domain), headers=headers)
    response = json.loads(r.text)
    if "error" in response:
        result = response["error"]["message"]
    else:
        stats = response["data"]["attributes"]["last_analysis_stats"]
        if stats["malicious"] > 0:
            result = "MALICIOUS"
        elif stats["suspicious"] > 0:
            result = "SUSPICIOUS"
        else:
            result = "CLEAN"
    ws.cell(row=i, column=1, value=domain)
    ws.cell(row=i, column=2, value=result)

# Save the output to an Excel file
wb.save(output_file)
