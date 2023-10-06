import requests
import json
import openpyxl
# VirusTotal API key
# apiKey = "6d20f0d4c9d8b0859dfea71e5937c6372e72813a617e10c582ad78215a74a017"

apifile = open('api.txt', 'r')
apiKey = apifile.readline()

# Input file path
inputFile = "input.txt"

# Output file path
outputFile = "output.xlsx"

# Create Excel workbook and sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set column headers
sheet["A1"] = "SHA256 Hash"
sheet["B1"] = "MD5 Hash"
sheet["C1"] = "SHA1 Hash"
# sheet["C1"] = "Detection Ratio"
# sheet["D1"] = "Scan Date"

# Read input file
with open(inputFile, "r") as file:
    hashes = file.readlines()

# Iterate through the list of SHA256 hashes
row = 2
for sha256 in hashes:
    sha256 = sha256.strip() # remove newline character

    # Construct API URL
    url = f"https://www.virustotal.com/api/v3/files/{sha256}"

    # Create API request headers
    headers = {
        "x-apikey": apiKey
    }

    # Send API request
    response = requests.get(url, headers=headers)

    # Parse response JSON
    data = json.loads(response.content.decode())

    # Extract MD5, detection information, and scan date
    try:
        md5 = data["data"]["attributes"]["md5"]
        sha1 = data["data"]["attributes"]["sha1"]
        detection = data["data"]["attributes"]["last_analysis_stats"]
        # detectionRatio = f"{detection['detected']}/{detection['total']}"
        if data["data"]["attributes"]["last_analysis_stats"]["malicious"] > 0:
            detect_result = "malicious"
        elif data["data"]["attributes"]["last_analysis_stats"]["suspicious"] > 0:
            detect_result = "Suspicious"
        else:
            detect_result = "Clean"
        scanDate = data["data"]["attributes"]["last_analysis_date"]

        # Write result to Excel
        sheet.cell(row, 1).value = sha256
        sheet.cell(row, 2).value = md5
        sheet.cell(row, 3).value = sha1
        #sheet.cell(row, 3).value = detectionRatio
        #sheet.cell(row, 4).value = scanDate
        sheet.cell(row, 5).value = detect_result

        # Increment row counter
        row += 1
        
    except KeyError:
        print("Unknown hash: {}".format(sha256))


# Save workbook
workbook.save(outputFile)

